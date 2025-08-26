#!/usr/bin/env python3
"""
run_me_nocerts.py

Credit-hour prep system (NO certificates yet).

What it does
------------
1) Cleans File A (Name, Hours, Event) for exact duplicate rows: (Name, Hours, Event).
2) Collapses File B (roster) to one canonical person per email (logs collisions).
3) Fuzzy-matches each EVENT ROW in File A to File B by name to retrieve Email/Category/etc.
4) Applies optional manual overrides (by FullName_A -> Email OR Full Name in File B).
5) Deduplicates by (Email, EventName) so a person can't overcount by using name variants.
6) Aggregates total hours by Email (canonical identity).
7) Optional Category filter (e.g., only "User").
8) Exports clean master list + diagnostics (Excel/CSV).

Required:
- Python 3.8+
- pandas
- openpyxl (writer/reader for .xlsx)

Usage
-----
python run_me_nocerts.py \
  --file_a "/path/to/FileA.xlsx" \
  --file_b "/path/to/FileB.xlsx" \
  --out_dir "/path/to/output_nocerts" \
  --category "User" \
  --min_match 0.88 \
  --overrides_csv "/path/to/manual_overrides.csv"

File A (first 3 columns, in order):   Full Name | Credit Hours | Event Name
File B (first 7 columns, in order):   Category | Subcategory | Full Name | Country | Email | CC Email | First Conference

Manual overrides CSV headers (flexible to case/spacing):
- FullName_A
- Override_FullName_B   (optional exact name from File B)
- Override_Email        (optional direct email; takes precedence when present)
"""


import argparse
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
from difflib import SequenceMatcher
import sys
import os

# GUI imports (lazy-used when --gui flag is set)
import threading
import subprocess
import platform
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except Exception:
    tk = None
    ttk = None
    filedialog = None
    messagebox = None


# ---------------------------
# Name normalization + fuzzy
# ---------------------------

_NICK_MAP = {
    "jon": "john",
    "johnathan": "jonathan",
    "pat": "patricia",
    "mike": "michael",
    "liz": "elizabeth",
    "beth": "elizabeth",
    "alex": "alexander",
    "sasha": "alexander",
    # Add more as desired...
}

def normalize_name(name: str) -> str:
    if not isinstance(name, str):
        return ""
    import re
    s = name.strip().lower()
    s = re.sub(r"[^a-z\s]", "", s)
    s = re.sub(r"\s+", " ", s)
    parts = s.split(" ")
    parts = [_NICK_MAP.get(p, p) for p in parts if p]
    return " ".join(parts)

def name_similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize_name(a), normalize_name(b)).ratio()


# --- Enhanced matching helpers ---
import re

def _strip_punct_and_spaces(s: str) -> str:
    if not isinstance(s, str):
        return ""
    return re.sub(r"[^a-z]", "", s.lower())

def _tokenize(name: str) -> list[str]:
    n = normalize_name(name)
    return [t for t in n.split(" ") if t]

def _initials(tokens: list[str]) -> str:
    return "".join(t[0] for t in tokens if t)

def composite_name_score(a: str, b: str) -> float:
    """Blend multiple signals to handle nicknames, hyphens, spacing, multi-last names, and FLIP order.
    Score in [0,1]."""
    a_norm = normalize_name(a)
    b_norm = normalize_name(b)

    # Base difflib similarity
    base = SequenceMatcher(None, a_norm, b_norm).ratio()

    # Token Jaccard (order-insensitive)
    at = set(_tokenize(a))
    bt = set(_tokenize(b))
    jacc = (len(at & bt) / len(at | bt)) if (at or bt) else 0.0

    # Initials boost (helps swapped order / middle names)
    ai = _initials(list(at))
    bi = _initials(list(bt))
    init = 1.0 if ai and bi and ai == bi else 0.0

    # Spacing/punctuation-insensitive similarity
    a_flat = _strip_punct_and_spaces(a)
    b_flat = _strip_punct_and_spaces(b)
    flat = SequenceMatcher(None, a_flat, b_flat).ratio() if a_flat and b_flat else 0.0

    # Weighted blend
    score = 0.45*base + 0.35*jacc + 0.10*flat + 0.10*init
    return max(0.0, min(1.0, score))


# --- Absolute name match helper ---
def is_absolute_name_match(a: str, b: str) -> bool:
    """Return True when names are an exact logical match after normalization.
    Treat spacing/punctuation differences and token order as non-issues.
    """
    na = normalize_name(a)
    nb = normalize_name(b)
    if na and nb and na == nb:
        return True
    # Flat (letters-only) equality
    if _strip_punct_and_spaces(a) == _strip_punct_and_spaces(b):
        return True
    # Same token bag (handles flipped order / middle names)
    if set(_tokenize(a)) == set(_tokenize(b)):
        return True
    # Handle concatenated names where File A has no spaces and File B has tokens possibly in any order
    # e.g., A: "Jangwanjae" vs B: "Wan Jae Jang" -> permutation concat equals A
    a_flat = _strip_punct_and_spaces(a)
    b_tokens = _tokenize(b)
    if 2 <= len(b_tokens) <= 4:  # avoid combinatorial blow-up
        try:
            from itertools import permutations
        except Exception:
            permutations = None
        if permutations:
            for perm in permutations(b_tokens):
                cand = "".join(_strip_punct_and_spaces(t) for t in perm)
                if a_flat == cand:
                    return True
    return False

# Strict equality ignoring only spacing/punctuation (no token reordering)
def is_spacing_punct_equal(a: str, b: str) -> bool:
    return _strip_punct_and_spaces(a) == _strip_punct_and_spaces(b)

def top_k_matches(name_a: str, df_b: pd.DataFrame, k: int = 3) -> list[tuple[int, float]]:
    """Return up to top-k matches as (row_index, score).
    If any spacing/punctuation/order-insensitive absolute match exists, ensure it is Top1 with score 1.0.
    """
    exact_hits = []
    fallback_scores = []
    b_names = df_b["Full Name"].fillna("")
    # First pass: collect absolute matches (regardless of spacing/punctuation/order)
    for i, name_b in enumerate(b_names):
        if is_absolute_name_match(name_a, name_b):
            exact_hits.append(i)
    if exact_hits:
        # Prefer an exact hit that also has an email; otherwise take the first exact hit
        exact_with_email = [i for i in exact_hits if str(df_b.iloc[i].get("Email", "")).strip() != ""]
        idx = exact_with_email[0] if exact_with_email else exact_hits[0]
        # Build the rest of the list from composite scores excluding the chosen index
        for j, name_b in enumerate(b_names):
            if j == idx:
                continue
            s = composite_name_score(name_a, name_b)
            fallback_scores.append((j, s))
        fallback_scores.sort(key=lambda x: x[1], reverse=True)
        return [(idx, 1.0)] + fallback_scores[: max(0, k - 1)]
    # No absolute hit: rank by composite score
    for i, name_b in enumerate(b_names):
        s = composite_name_score(name_a, name_b)
        fallback_scores.append((i, s))
    fallback_scores.sort(key=lambda x: x[1], reverse=True)
    return fallback_scores[:k]


# ---------------------------
# Cleaning + helpers
# ---------------------------

def dedupe_exact_file_a(df_a: pd.DataFrame) -> pd.DataFrame:
    """Remove exact duplicates of (Name, Hours, Event)."""
    temp = df_a.copy()
    temp.columns = ["FullName_A", "CreditHours", "EventName"]
    key = (
        temp["FullName_A"].astype(str).str.strip().str.lower() + "||" +
        temp["CreditHours"].astype(str) + "||" +
        temp["EventName"].astype(str).str.strip().str.lower()
    )
    temp["__key"] = key
    temp = temp.drop_duplicates(subset="__key").drop(columns="__key")
    return temp

def collapse_roster_by_email(df_b_raw: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Collapse File B to one row per unique (non-empty) Email.
    Returns (collapsed_df, email_collisions_df).
    """
    dfb = df_b_raw.copy()
    dfb.columns = ["Category", "Subcategory", "Full Name", "Country", "Email", "CC Email", "First Conference"]

    # Detect collisions where the same email appears multiple times
    non_empty = dfb["Email"].astype(str).str.strip() != ""
    collisions = (
        dfb[non_empty].groupby("Email").size().reset_index(name="Count")
    )
    collisions = collisions[collisions["Count"] > 1]

    # Keep first row for each email; keep rows with empty emails unchanged
    dfb_email = dfb[non_empty].copy()
    dfb_noemail = dfb[~non_empty].copy()
    dfb_email = dfb_email.sort_values(["Email"]).drop_duplicates(subset=["Email"], keep="first")
    collapsed = pd.concat([dfb_email, dfb_noemail], ignore_index=True)

    return collapsed, collisions

def fuzzy_match_name_to_b_row(name_a: str, df_b: pd.DataFrame, min_score: float) -> tuple[Optional[pd.Series], float]:
    best = top_k_matches(name_a, df_b, k=1)
    if not best:
        return None, 0.0
    idx, score = best[0]
    if score < float(min_score):
        # Still return the best row, but the caller can mark as low-confidence
        return df_b.iloc[idx], float(score)
    return df_b.iloc[idx], float(score)

def load_overrides(overrides_csv: str) -> Optional[pd.DataFrame]:
    if not overrides_csv:
        return None
    p = Path(overrides_csv)
    if not p.exists():
        return None
    ov = pd.read_csv(p).fillna("")
    # Flexible header matching
    mapping = {c.lower().strip(): c for c in ov.columns}
    fn_col = mapping.get("fullname_a") or mapping.get("full_name_a") or "FullName_A"
    ob_col = mapping.get("override_fullname_b") or "Override_FullName_B"
    oe_col = mapping.get("override_email") or "Override_Email"

    for col in [fn_col, ob_col, oe_col]:
        if col not in ov.columns:
            ov[col] = ""

    ov = ov[[fn_col, ob_col, oe_col]].copy()
    ov.columns = ["FullName_A", "Override_FullName_B", "Override_Email"]
    return ov

def apply_overrides_event_level(
    df_joined_events: pd.DataFrame,
    df_b: pd.DataFrame,
    overrides_df: Optional[pd.DataFrame]
) -> pd.DataFrame:
    """Apply manual overrides per event row."""
    if overrides_df is None or overrides_df.empty:
        return df_joined_events

    # Quick lookup by full name from B
    bcols = ["Full Name", "Category", "Subcategory", "Country", "Email", "CC Email", "First Conference"]
    b_lookup: Dict[str, Dict[str, str]] = df_b.set_index("Full Name")[bcols].to_dict(orient="index") # type: ignore

    df = df_joined_events.merge(overrides_df, on="FullName_A", how="left")
    # Fill missing cols
    if "Override_FullName_B" not in df.columns:
        df["Override_FullName_B"] = ""
    if "Override_Email" not in df.columns:
        df["Override_Email"] = ""

    # Apply overrides
    for idx, row in df.iterrows():
        override_email = str(row.get("Override_Email", "")).strip()
        override_bname = str(row.get("Override_FullName_B", "")).strip()

        if override_email:
            df.at[idx, "Email"] = override_email
            if not str(row.get("MatchedName_B", "")).strip():
                df.at[idx, "MatchedName_B"] = row.get("FullName_A", "")
            df.at[idx, "MatchSource"] = "OVERRIDDEN_EMAIL"
            df.at[idx, "MatchScore"] = 1.0
            df.at[idx, "ReviewFlag"] = ""
            continue

        if override_bname:
            if override_bname in b_lookup:
                rec = b_lookup[override_bname]
                df.at[idx, "MatchedName_B"] = override_bname
                df.at[idx, "Email"] = rec.get("Email", "")
                df.at[idx, "Category"] = rec.get("Category", "")
                df.at[idx, "Subcategory"] = rec.get("Subcategory", "")
                df.at[idx, "Country"] = rec.get("Country", "")
                df.at[idx, "CC Email"] = rec.get("CC Email", "")
                df.at[idx, "First Conference"] = rec.get("First Conference", "")
                df.at[idx, "MatchSource"] = "OVERRIDDEN_NAME"
                df.at[idx, "MatchScore"] = 1.0
                df.at[idx, "ReviewFlag"] = ""

    # Drop helper override columns before returning
    return df.drop(columns=[c for c in ["Override_FullName_B", "Override_Email"] if c in df.columns])



# ---------------------------
# Main pipeline
def _open_path(path: str):
    try:
        if platform.system() == "Darwin":
            subprocess.Popen(["open", path])
        elif platform.system() == "Windows":
            os.startfile(path)  # type: ignore
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass
def launch_gui():
    if tk is None:
        print("Tkinter not available in this environment. Install tkinter and try again.")
        return

    root = tk.Tk()
    root.title("Credit Hours Prep (No Certificates)")
    root.geometry("720x520")

    # Vars
    file_a_var = tk.StringVar()
    file_b_var = tk.StringVar()
    out_dir_var = tk.StringVar(value=str(Path.home() / "Desktop" / "output_nocerts"))
    category_var = tk.StringVar()
    min_match_var = tk.StringVar(value="0.85")
    decisions_var = tk.StringVar()
    overrides_var = tk.StringVar()

    def pick_file_a():
        p = filedialog.askopenfilename(title="Select File A (Name, Hours, Event)", filetypes=[("Excel", ".xlsx .xls")])
        if p:
            file_a_var.set(p)

    def pick_file_b():
        p = filedialog.askopenfilename(title="Select File B (Roster)", filetypes=[("Excel", ".xlsx .xls")])
        if p:
            file_b_var.set(p)

    def pick_out_dir():
        p = filedialog.askdirectory(title="Select Output Folder")
        if p:
            out_dir_var.set(p)

    def pick_decisions():
        p = filedialog.askopenfilename(title="Select decisions file (proposed_matches.xlsx or CSV)", filetypes=[("Excel/CSV", ".xlsx .csv")])
        if p:
            decisions_var.set(p)

    def pick_overrides():
        p = filedialog.askopenfilename(title="Select manual_overrides.csv", filetypes=[("CSV", ".csv")])
        if p:
            overrides_var.set(p)

    # Layout helpers
    def row(label, var, picker=None):
        frame = ttk.Frame(main)
        frame.pack(fill="x", padx=8, pady=4)
        ttk.Label(frame, text=label, width=22).pack(side="left")
        e = ttk.Entry(frame, textvariable=var)
        e.pack(side="left", fill="x", expand=True)
        if picker:
            ttk.Button(frame, text="Browse", command=picker).pack(side="left", padx=6)
        return e

    def run_in_thread(fn):
        def _wrapped():
            btn_disable()
            log_var.set("Running... this may take a moment.\n")
            def target():
                try:
                    fn()
                    log_var.set(log_var.get() + "\nDone.\n")
                except Exception as e:
                    log_var.set(log_var.get() + f"\nError: {e}\n")
                    try:
                        messagebox.showerror("Error", str(e))
                    except Exception:
                        pass
                finally:
                    btn_enable()
            threading.Thread(target=target, daemon=True).start()
        return _wrapped

    def btn_disable():
        for b in buttons:
            b.config(state="disabled")

    def btn_enable():
        for b in buttons:
            b.config(state="normal")

    def do_generate_proposals():
        run_pipeline(
            file_a=file_a_var.get(),
            file_b=file_b_var.get(),
            out_dir=out_dir_var.get(),
            min_match=float(min_match_var.get() or 0.85),
            category_filter=(category_var.get() or None),
            overrides_csv=(overrides_var.get() or None),
            decisions_path=None,
        )
        # Open proposals file if present
        pm = Path(out_dir_var.get())/"proposed_matches.xlsx"
        if pm.exists():
            _open_path(str(pm))
        _open_path(out_dir_var.get())

    def do_apply_decisions():
        # Default to proposed_matches.xlsx in output if no path provided
        dec_path = decisions_var.get() or str(Path(out_dir_var.get())/"proposed_matches.xlsx")
        run_pipeline(
            file_a=file_a_var.get(),
            file_b=file_b_var.get(),
            out_dir=out_dir_var.get(),
            min_match=float(min_match_var.get() or 0.85),
            category_filter=(category_var.get() or None),
            overrides_csv=(overrides_var.get() or None),
            decisions_path=dec_path,
        )
        # Open master list when done
        ml = Path(out_dir_var.get())/"master_list.xlsx"
        if ml.exists():
            _open_path(str(ml))
        _open_path(out_dir_var.get())

    # UI
    main = ttk.Frame(root)
    main.pack(fill="both", expand=True)

    ttk.Label(main, text="Credit Hours Prep (No Certificates)", font=("TkDefaultFont", 14, "bold")).pack(pady=(10,4))
    ttk.Label(main, text="Step 1: Select inputs → Generate Proposals.\nStep 2: Review & mark ACCEPT in proposed_matches.xlsx.\nStep 3: Apply Decisions to build the master list.", foreground="#555").pack()

    row("File A (PDH Data):", file_a_var, pick_file_a)
    row("File B (Registration):", file_b_var, pick_file_b)
    row("Output Folder:", out_dir_var, pick_out_dir)
    row("Decisions File (opt):", decisions_var, pick_decisions)
    row("Overrides CSV (opt):", overrides_var, pick_overrides)

    # Options row
    opts = ttk.Frame(main)
    opts.pack(fill="x", padx=8, pady=6)
    ttk.Label(opts, text="Category (opt):").pack(side="left")
    ttk.Entry(opts, textvariable=category_var, width=18).pack(side="left", padx=(4,10))
    ttk.Label(opts, text="Min Match (0-1):").pack(side="left")
    ttk.Entry(opts, textvariable=min_match_var, width=8).pack(side="left", padx=(4,10))

    # Buttons
    btns = ttk.Frame(main)
    btns.pack(fill="x", padx=8, pady=8)
    b1 = ttk.Button(btns, text="Generate Proposals", command=run_in_thread(do_generate_proposals))
    b2 = ttk.Button(btns, text="Apply Decisions", command=run_in_thread(do_apply_decisions))
    b3 = ttk.Button(btns, text="Open Output Folder", command=lambda: _open_path(out_dir_var.get()))
    buttons = [b1, b2, b3]
    for b in buttons:
        b.pack(side="left", padx=6)

    # Log area
    log_var = tk.StringVar(value="\nReady.\n")
    log = tk.Text(main, height=12)
    log.pack(fill="both", expand=True, padx=8, pady=8)

    def sync_log(*_):
        log.delete("1.0", "end")
        log.insert("end", log_var.get())
        log.see("end")
    log_var.trace_add("write", sync_log)
    sync_log()

    root.mainloop()

def run_pipeline(
    file_a: str,
    file_b: str,
    out_dir: str,
    min_match: float = 0.85,
    category_filter: Optional[str] = None,
    overrides_csv: Optional[str] = None,
    decisions_path: Optional[str] = None,
) -> None:

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    # Load inputs and standardize columns by position
    df_a_raw = pd.read_excel(file_a, header=0)
    df_a = df_a_raw.iloc[:, :3].copy()
    df_a.columns = ["FullName_A", "CreditHours", "EventName"]
    df_a = df_a.dropna(subset=["FullName_A", "CreditHours"])

    df_b_raw = pd.read_excel(file_b, header=0)
    df_b_raw = df_b_raw.iloc[:, :7].copy()
    df_b_raw.columns = ["Category", "Subcategory", "Full Name", "Country", "Email", "CC Email", "First Conference"]

    # Collapse roster to one row per email and log collisions
    df_b, email_collisions = collapse_roster_by_email(df_b_raw)
    if not email_collisions.empty:
        email_collisions.to_excel(out_path / "roster_email_duplicates.xlsx", index=False)

    # Step 1: Clean blatant duplicates in A
    df_a_clean = dedupe_exact_file_a(df_a)

    # Build proposed matches (per unique name in A) to aid manual review — SINGLE FILE
    unique_names = sorted(set(df_a_clean["FullName_A"].astype(str)))
    proposal_rows = []
    for nm in unique_names:
        tops = top_k_matches(nm, df_b, k=3)
        entry = {"FullName_A": nm}
        if tops:
            idx1, sc1 = tops[0]
            br1 = df_b.iloc[idx1]
            entry["Top1_Name_B"] = br1.get("Full Name", "")
            entry["Top1_Email"] = br1.get("Email", "")
            entry["Top1_Score"] = round(float(sc1), 3)
            # Decide "Certain" only when exact letters-only equality (spacing/punctuation-insensitive) with email present
            is_certain = (entry.get("Top1_Score", 0.0) == 1.0) and is_spacing_punct_equal(nm, entry["Top1_Name_B"]) and bool(entry["Top1_Email"])
            entry["Certain"] = bool(is_certain)

            # Include Top2/Top3 only if NOT certain (to reduce clutter)
            if not is_certain and len(tops) > 1:
                idx2, sc2 = tops[1]
                br2 = df_b.iloc[idx2]
                entry["Top2_Name_B"] = br2.get("Full Name", "")
                entry["Top2_Email"] = br2.get("Email", "")
                entry["Top2_Score"] = round(float(sc2), 3)
            else:
                entry["Top2_Name_B"], entry["Top2_Email"], entry["Top2_Score"] = "", "", ""
            if not is_certain and len(tops) > 2:
                idx3, sc3 = tops[2]
                br3 = df_b.iloc[idx3]
                entry["Top3_Name_B"] = br3.get("Full Name", "")
                entry["Top3_Email"] = br3.get("Email", "")
                entry["Top3_Score"] = round(float(sc3), 3)
            else:
                entry["Top3_Name_B"], entry["Top3_Email"], entry["Top3_Score"] = "", "", ""

            # Suggested + decision fields
            entry["Suggested_Email"] = entry.get("Top1_Email", "")
            # Auto-accept only when Certain; otherwise leave blank for human review
            entry["Decision"] = "ACCEPT" if is_certain else ""
            entry["Chosen_Email"] = ""
        else:
            # No candidates — force review
            entry.update({
                "Top1_Name_B": "", "Top1_Email": "", "Top1_Score": "",
                "Top2_Name_B": "", "Top2_Email": "", "Top2_Score": "",
                "Top3_Name_B": "", "Top3_Email": "", "Top3_Score": "",
                "Certain": False, "Suggested_Email": "", "Decision": "", "Chosen_Email": ""
            })
        proposal_rows.append(entry)

    proposals_df = pd.DataFrame(proposal_rows)
    # Sort so items needing attention appear first and greens last
    proposals_df = proposals_df.sort_values(by=["Certain", "Top1_Score", "FullName_A"], ascending=[True, True, True])

    # Write ONE Excel (no conditional formatting for auto-certain rows)
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Proposed Matches"

    for r in dataframe_to_rows(proposals_df, index=False, header=True):
        ws.append(r)

    # Green highlight for rows where Certain == TRUE
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import FormulaRule

    header = [cell.value for cell in ws[1]]
    if "Certain" in header:
        certain_col_idx = header.index("Certain") + 1
        last_row = ws.max_row
        last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
        certain_col_letter = ws.cell(row=1, column=certain_col_idx).column_letter
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        rule = FormulaRule(formula=[f"${certain_col_letter}2=TRUE"], stopIfTrue=False, fill=green_fill)
        ws.conditional_formatting.add(f"A2:{last_col_letter}{last_row}", rule)

    pm_path = Path(out_dir)/"proposed_matches.xlsx"
    wb.save(pm_path)

    # Step 2: Event-level fuzzy matching
    joined_rows = []
    for _, r in df_a_clean.iterrows():
        name_a = r["FullName_A"]
        hours = float(r["CreditHours"])
        event = r["EventName"]
        match_row, score = fuzzy_match_name_to_b_row(name_a, df_b, min_match)

        if match_row is None:
            joined_rows.append({
                "FullName_A": name_a,
                "EventName": event,
                "CreditHours": hours,
                "MatchedName_B": "",
                "Email": "",
                "Category": "",
                "Subcategory": "",
                "Country": "",
                "CC Email": "",
                "First Conference": "",
                "MatchScore": round(float(score), 3) if score else 0.0,
                "MatchSource": "FUZZY_NO_MATCH",
                "ReviewFlag": "NO_MATCH_OR_LOW_SCORE",
                "Confidence": round(float(score), 3) if score else 0.0,
                "AssumedEmail": "",
                "AssumedName": "",
            })
        else:
            joined_rows.append({
                "FullName_A": name_a,
                "EventName": event,
                "CreditHours": hours,
                "MatchedName_B": match_row["Full Name"],
                "Email": match_row["Email"],
                "Category": match_row["Category"],
                "Subcategory": match_row["Subcategory"],
                "Country": match_row["Country"],
                "CC Email": match_row["CC Email"],
                "First Conference": match_row["First Conference"],
                "MatchScore": round(float(score), 3),
                "MatchSource": "FUZZY_NAME",
                "ReviewFlag": "",
                "Confidence": round(float(score), 3),
                "AssumedEmail": match_row["Email"],
                "AssumedName": match_row["Full Name"],
            })

    df_joined_events = pd.DataFrame(joined_rows)
    df_joined_events.to_excel(out_path / "joined_events_pre_overrides.xlsx", index=False)

    # Optional: apply decisions from the single proposed_matches file (xlsx or csv)
    if decisions_path is None:
        default_pm = Path(out_dir)/"proposed_matches.xlsx"
        decisions_path = str(default_pm) if default_pm.exists() else None

    if decisions_path:
        dec_path = Path(decisions_path)
        if dec_path.exists():
            if dec_path.suffix.lower() == ".xlsx":
                dec = pd.read_excel(dec_path).fillna("")
            else:
                dec = pd.read_csv(dec_path).fillna("")
            # Normalize headers
            cmap = {c.lower().strip(): c for c in dec.columns}
            fnc = cmap.get("fullname_a") or "FullName_A"
            sugg = cmap.get("suggested_email") or "Suggested_Email"
            deci = cmap.get("decision") or "Decision"
            chosen = cmap.get("chosen_email") or "Chosen_Email"
            for c in (fnc, sugg, deci, chosen):
                if c not in dec.columns:
                    dec[c] = ""
            dec = dec[[fnc, sugg, deci, chosen]].copy()
            dec.columns = ["FullName_A", "Suggested_Email", "Decision", "Chosen_Email"]

            df_joined_events = df_joined_events.merge(dec, on="FullName_A", how="left")
            # Apply decisions (ACCEPT/REJECT) with optional Chosen_Email
            for idx, row in df_joined_events.iterrows():
                decision = str(row.get("Decision", "")).strip().upper()
                chosen_email = str(row.get("Chosen_Email", "")).strip()
                suggested_email = str(row.get("Suggested_Email", "")).strip()
                if decision == "ACCEPT":
                    email_to_set = chosen_email or suggested_email
                    if email_to_set:
                        df_joined_events.at[idx, "Email"] = email_to_set
                        if not str(row.get("MatchedName_B", "")).strip():
                            df_joined_events.at[idx, "MatchedName_B"] = row.get("FullName_A", "")
                        df_joined_events.at[idx, "MatchSource"] = "USER_ACCEPTED"
                        df_joined_events.at[idx, "ReviewFlag"] = ""
                elif decision == "REJECT":
                    df_joined_events.at[idx, "Email"] = ""
                    df_joined_events.at[idx, "MatchSource"] = "USER_REJECTED"
                    df_joined_events.at[idx, "ReviewFlag"] = "REJECTED_NEEDS_EMAIL"

    # Mark low-confidence auto-assignments
    low_mask = (df_joined_events["Confidence"].fillna(0) < float(min_match)) & (df_joined_events["Email"].astype(str).str.strip() != "")
    df_joined_events.loc[low_mask, "ReviewFlag"] = df_joined_events.loc[low_mask, "ReviewFlag"].replace("", "LOW_CONFIDENCE_AUTOASSIGN")

    # Step 3: Apply manual overrides
    overrides_df = load_overrides(overrides_csv) if overrides_csv else None
    df_joined_events = apply_overrides_event_level(df_joined_events, df_b, overrides_df)
    df_joined_events.to_excel(out_path / "joined_events.xlsx", index=False)

    # Step 4: Deduplicate by (Email, EventName) to prevent overcounting
    df_with_email = df_joined_events[df_joined_events["Email"].astype(str).str.strip() != ""].copy()
    df_without_email = df_joined_events[df_joined_events["Email"].astype(str).str.strip() == ""].copy()

    # Prefer higher MatchScore for duplicates with same (Email, EventName)
    df_with_email = df_with_email.sort_values(by=["Email", "EventName", "MatchScore"], ascending=[True, True, False])

    # Mark removed duplicates (same Email+Event)
    dup_mask = df_with_email.duplicated(subset=["Email", "EventName"], keep="first")
    removed_dups = df_with_email[dup_mask].copy()
    if not removed_dups.empty:
        removed_dups.to_excel(out_path / "duplicates_removed_same_email_event.xlsx", index=False)

    df_with_email_dedup = df_with_email[~dup_mask].copy()

    # Step 5: Aggregate hours by Email
    # Log unmatched (no email) for manual fix
    if not df_without_email.empty:
        df_without_email.to_excel(out_path / "unmatched_needs_email.xlsx", index=False)

    # Pick a canonical display name per email (from collapsed File B), else fallback
    canonical_name_map = (
        df_b[["Email", "Full Name"]].drop_duplicates().set_index("Email")["Full Name"].to_dict()
    )
    df_with_email_dedup["DisplayName"] = (
        df_with_email_dedup["Email"].map(canonical_name_map)
        .fillna(df_with_email_dedup["MatchedName_B"])
        .fillna(df_with_email_dedup["FullName_A"])
    )

    totals = df_with_email_dedup.groupby(["Email"], as_index=False).agg({
        "DisplayName": "first",
        "Category": "first",
        "Subcategory": "first",
        "Country": "first",
        "First Conference": "first",
        "CreditHours": "sum",
    })
    totals["TotalCreditHours"] = totals["CreditHours"].astype(float).round(2)
    totals = totals.drop(columns=["CreditHours"])

    # Step 6: Optional Category filter
    if category_filter:
        mask = totals["Category"].astype(str).str.strip().str.lower() == category_filter.strip().lower()
        excluded = totals[~mask].copy()
        if not excluded.empty:
            excluded.to_excel(out_path / "excluded_by_category.xlsx", index=False)
        totals = totals[mask].copy()

    # Step 7: Exports
    master_cols = ["DisplayName", "Email", "TotalCreditHours", "Category", "Subcategory"]
    totals[master_cols].to_excel(out_path / "master_list.xlsx", index=False)
    totals[master_cols].to_csv(out_path / "master_list.csv", index=False)

    # Extra: quick audit table
    audit_cols = ["FullName_A", "MatchedName_B", "Email", "EventName", "CreditHours", "MatchScore", "MatchSource"]
    df_with_email_dedup[audit_cols].to_excel(out_path / "event_level_audit.xlsx", index=False)

    print(f"Done. Outputs in: {out_path.resolve()}")


def main():
    # If launched with no arguments (e.g., double-clicked app), open GUI by default
    if len(sys.argv) == 1:
        launch_gui()
        return
    parser = argparse.ArgumentParser(
        description="Prepare credit-hour master list (no certificates) with email-based canonicalization."
    )
    parser.add_argument("--file_a", required=False, help="Path to File A (Name, Hours, Event)")
    parser.add_argument("--file_b", required=False, help="Path to File B (Category, Subcategory, Full Name, Country, Email, CC Email, First Conference)")
    parser.add_argument("--out_dir", default="output_nocerts", help="Output directory path")
    parser.add_argument("--category", required=False, help="Optional Category filter (e.g., 'User')")
    parser.add_argument("--min_match", type=float, default=0.85, help="Minimum fuzzy match score (0-1)")
    parser.add_argument("--overrides_csv", required=False, help="Path to manual_overrides.csv (optional)")
    parser.add_argument("--decisions_path", required=False, help="Path to decisions file (use proposed_matches.xlsx or a CSV). Optional.")
    parser.add_argument("--gui", action="store_true", help="Launch graphical app instead of CLI")
    args = parser.parse_args()

    # If not using GUI, require the two input files
    if not args.gui:
        if not args.file_a or not args.file_b:
            parser.error("the following arguments are required in CLI mode: --file_a, --file_b")

    if args.gui:
        launch_gui()
        return

    run_pipeline(
        file_a=args.file_a,
        file_b=args.file_b,
        out_dir=args.out_dir,
        min_match=args.min_match,
        category_filter=args.category,
        overrides_csv=args.overrides_csv,
        decisions_path=args.decisions_path,
    )


if __name__ == "__main__":
    main()
